from openpyxl import load_workbook,Workbook
import datetime,time
import csv,re
from collections import OrderedDict

formula_file="Formula.xlsx"

localtime = time.strftime("%Y%m%d", time.localtime())

filepath_CC1 = ".\PV\CC1\\"
filepath_CC2 = ".\PV\CC2\\"
filepath_CC3 = ".\PV\CC3\\"
filepath_CC4 = ".\PV\CC4\\"
filepath_CC5 = ".\PV\CC5\\"

cc1_filepath = filepath_CC1+localtime+".csv"


cc1_filepath = ".\PV\CC1\\20170514.csv"
cc2_filepath = ".\PV\CC2\\20170514.csv"
cc3_filepath = ".\PV\CC3\\20170514.csv"
cc4_filepath = ".\PV\CC4\\20170514.csv"
cc5_filepath = ".\PV\CC5\\20170514.csv"

serial_row_n=5
yeild_row_n=6
kWh_row_n=8
start_time_row_n=9

def process_cc(input_filename,cc_serial_dict):
    print ("Processing file",input_filename)
    start_time=datetime.datetime.now()
    new_filepath=input_filename.rstrip(".csv")+"_output.csv"

    csvfile = open(input_filename, 'r')  #changed by XY, from 'rb' to 'r',20170918
    reader = csv.reader(csvfile)
    rows_list=[]
    for row in reader:
       rr=re.split(';',''.join(row))
       rows_list.append(list(rr))
       
    serial_row=rows_list[serial_row_n]
    kWh_row=rows_list[kWh_row_n]
    yeild_row=rows_list[yeild_row_n]
    
    csv_write_file = open(new_filepath, 'w')   #changed by XY, from 'wb' to 'w',20170918
    writer = csv.writer(csv_write_file, delimiter=',')#,quotechar='|', quoting=csv.QUOTE_MINIMAL)
    
    def get_kwh_index(serial):
        for i in range(len(serial_row)):
            
            if(serial_row[i]==serial): 
               if(kWh_row[i]=="kWh"):# and yeild_row[i] =="Total yield"):           
                   return i
        return -1
 
    def process():
        name_kwd_ind=OrderedDict()
        for name,serials in cc_serial_dict.items():
            kwh_ind=list()
            for serial in serials:
                index=get_kwh_index(str(serial))
                if(index!=-1):
                   kwh_ind.append(index)
                else:
                   print ("Serial ",serial," not found")
                   time.sleep(1)
               
            name_kwd_ind[name]=list(kwh_ind)
        #print ("name",list(name_kwd_ind.keys()))
        writer.writerow(["From Timestamp"] + list(name_kwd_ind.keys())) #changed by BXY, use functuion list 20170918
        for i in range(start_time_row_n,len(rows_list)):
           
                sum=OrderedDict()
                sums=[]
                sums.append(rows_list[i][0])
                for n,l in name_kwd_ind.items():
                    sum[n]=0
                    for kwh_ind in l:
                        sum[n]+=float(rows_list[i][kwh_ind])
                    
                strng=""
                row
                for n,s in sum.items():
                   strng+=n+":"+str(s)+"|"
                   sums.append(s)
                writer.writerow(sums)
				
    #print "Total number of rows in ",input_filename,":",len(rows_list)
    process()
    csvfile.close()
    csv_write_file.close()
    
def get_formula_CC_serials(file_name):
    print ("Processing formula file")
    wb = load_workbook(filename=file_name, read_only=True)
    ws = wb.active
    cc1_serials=OrderedDict()
    cc2_serials=OrderedDict()
    cc3_serials=OrderedDict()
    cc4_serials=OrderedDict()
    cc5_serials=OrderedDict()
    
    for formula_row in ws.iter_rows(row_offset=3):
        if(formula_row[1].value):
            if(formula_row[1].value not in cc1_serials.keys()):
               cc1_serials[formula_row[1].value]=list()    
            cc1_serials[formula_row[1].value].append(formula_row[3].value)
        if(formula_row[5].value):
            if(formula_row[5].value not in cc2_serials.keys()):
               cc2_serials[formula_row[5].value]=list()    
            cc2_serials[formula_row[5].value].append(formula_row[7].value)
        if(formula_row[9].value):
            if(formula_row[9].value not in cc3_serials.keys()):
               cc3_serials[formula_row[9].value]=list()    
            cc3_serials[formula_row[9].value].append(formula_row[11].value)
        if(formula_row[13].value):
            if(formula_row[13].value not in cc4_serials.keys()):
               cc4_serials[formula_row[13].value]=list()    
            cc4_serials[formula_row[13].value].append(formula_row[15].value)
        if(formula_row[17].value):
            if(formula_row[17].value not in cc5_serials.keys()):
               cc5_serials[formula_row[17].value]=list()    
            cc5_serials[formula_row[17].value].append(formula_row[19].value)
        
    return cc1_serials,cc2_serials,cc3_serials,cc4_serials,cc5_serials    


if __name__ == '__main__': 
    print ("Statred:",str(datetime.datetime.now()))
    start_time=datetime.datetime.now()
    
    #get CC formula serials  from formula file
    cc1_serials,cc2_serials,cc3_serials,cc4_serials,cc5_serials  = get_formula_CC_serials(formula_file)
    
    #process CC1
    process_cc(cc1_filepath,cc1_serials)
    #process CC2
    process_cc(cc2_filepath,cc2_serials)
    #process CC3
    process_cc(cc3_filepath,cc3_serials)
    #process CC4
    process_cc(cc4_filepath,cc4_serials)
    #process CC5
    process_cc(cc5_filepath,cc5_serials)
    

print ("Finished:", str(datetime.datetime.now()))
print ("Time elapsed:", str(datetime.datetime.now()-start_time))
